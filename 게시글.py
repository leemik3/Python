import time
from selenium import webdriver
import openpyxl
from selenium.webdriver import ActionChains
from selenium.common.exceptions import NoSuchElementException
from urllib.request import urlretrieve



try:
    wb1 = openpyxl.load_workbook("crawlingpost.xlsx")
    wb2 = openpyxl.load_workbook("crawlingpostimage.xlsx")
except:
    wb1 = openpyxl.Workbook()
    wb2 = openpyxl.Workbook()
    sheet1 = wb1.active
    sheet1.append(["id","team", "title","writer","content"])
    sheet2 = wb2.active
    sheet2.append(["postid","url"])
else:
    sheet1 = wb1.active
    sheet2 = wb2.active

driver = webdriver.Chrome("./chromedriver")

driver.get("http://cafe118.daum.net/_c21_/bbs_list?grpid=aVeZ&fldid=8mNz&listnum=20")
time.sleep(12)
driver.get("http://cafe118.daum.net/_c21_/bbs_read?grpid=aVeZ&fldid=8mNz&contentval=00Da2zzzzzzzzzzzzzzzzzzzzzzzzz&datanum=52206&page=1&prev_page=0&firstbbsdepth=&lastbbsdepth=zzzzzzzzzzzzzzzzzzzzzzzzzzzzzz&listnum=20")

id = 0

for i in range(300):
    try:
        team = driver.find_element_by_css_selector("strong.tit_info>span").text
    except NoSuchElementException:
        team = ""
    title = driver.find_element_by_css_selector("strong.tit_info").text[len(team):].strip()
    writer = driver.find_element_by_css_selector("a.link_item").text
    content = driver.find_element_by_css_selector("div#user_contents").text
    for j in ["씨발", "시발", "좆", "ㅅㅂ", "좃", "씨바", "ㅈㄹ", "ㅄ", "ㅂㅅ", "병신", "병시나", "새끼", "지랄", "ㅈㄹ", "엿", "자지", "야동", "애미","씹", "붕가"]:
        if content.__contains__(j) == True:
            continue

    id += 1
    sheet1.append([id, team, title, writer, content])

    image = driver.find_elements_by_css_selector("div#user_contents img")
    for i in range(len(image)):
        try:
            link = image[i].get_attribute("src")
        except:
            continue
        urlretrieve(link, "./크롤링이미지/"+str(id)+"_"+str(i)+".jpg")
        sheet2.append([id,link])

    driver.find_element_by_css_selector("a#after-article").click()
    time.sleep(1)
'''
driver.get("http://cafe.daum.net/dotax/LSNc/197295")
time.sleep(12)
driver.get("http://cafe.daum.net/dotax/LSNc/197295")

for i in range(300):
    try:
        team = driver.find_element_by_css_selector("strong.tit_info>span").text
    except NoSuchElementException:
        team = ""
    title = driver.find_element_by_css_selector("strong.tit_info").text.replace(team,"").strip()
    writer = driver.find_element_by_css_selector("a.link_item").text
    content = driver.find_element_by_css_selector("div#user_contents").text
    if ["씨발","시발","좆","ㅅㅂ","좃", "씨바", "ㅈㄹ", "ㅄ", "ㅂㅅ", "병신", "병시나", "새끼", "지랄", "ㅈㄹ", "엿", "자지", "야동", "애미", "씹", "붕가"] in content:
        continue

    id += 1
    sheet1.append([id, team, title, writer, content])

    image = driver.find_elements_by_css_selector("div#user_contents img")
    for i in range(len(image)):
        try:
            link = image[i].get_attribute("src")
        except:
            continue
        urlretrieve(link, "./크롤링이미지/"+str(id)+"_"+str(i)+".jpg")
        sheet2.append([id,link])

    driver.find_element_by_css_selector("a#after-article").click()
    time.sleep(1)
'''

wb1.save("crawlingpost.xlsx")
wb2.save("crawlingpostimage.xlsx")

'''
# qna찾기 위해 끝까지 스크롤

last_height = driver.execute_script("return document.documentElement.scrollHeight")
while True:
    print("스크롤", end="")
    driver.execute_script("window.scrollTo(0, document.documentElement.scrollHeight);")
    print(" 완료, 2초 대기")

    time.sleep(2)

    new_height = driver.execute_script("return document.documentElement.scrollHeight")

    if new_height == last_height:
        break
    last_height = new_height

cont = driver.find_element_by_css_selector("div#main")

comment = driver.find_elements_by_css_selector("div#contents div#content")

n = 0
for i in comment:
    n+=1
    try:  # 자세히보기 버튼 클릭 try
        p = i.find_element_by_xpath('../..')
        p.find_element_by_css_selector("span.more-button.style-scope.ytd-comment-renderer").click()
        id_ = p.find_element_by_css_selector("span.style-scope.ytd-comment-renderer").text
        com = i.find_element_by_css_selector("yt-formatted-string.style-scope#content-text").text
        print(com)
        sheet.append([n, id_, com])
    except:
        p = i.find_element_by_xpath('../..')
        id_ = p.find_element_by_css_selector("span.style-scope.ytd-comment-renderer").text
        com = i.find_element_by_css_selector("yt-formatted-string.style-scope#content-text").text
        print(com)
        sheet.append([n, id_, com])

wb.save("crawlingpost.xlsx")
wb2.save("crawlingpostimage.xlsx")


'''